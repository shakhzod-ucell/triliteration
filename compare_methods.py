"""
compare_methods.py
------------------
Generates side-by-side comparison of TA vs RSRP positioning methods.
Shows statistical summary, tower-by-tower comparison, and complete error lists.
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def generate_comparison_report(results: dict, output_path: str):
    """
    Generate Excel comparison report for multiple positioning methods.

    Parameters
    ----------
    results : dict
        Keys are method names ('TA', 'RSRP'), values are dicts with:
        - 'matched': DataFrame with error_m column
        - 'towers': DataFrame with positioning results
        - 'strategy': distance strategy object
    output_path : str
        Path to save Excel file
    """

    # ── Extract data ─────────────────────────────────────────────────────
    methods = list(results.keys())
    if len(methods) < 2:
        print(f"  ⚠️  Need at least 2 methods for comparison, got {len(methods)}")
        return

    # ── Sheet 1: Statistical Summary ─────────────────────────────────────
    summary_data = []
    for method_name in methods:
        errors = results[method_name]['matched']['error_m']
        strategy = results[method_name]['strategy']

        summary_data.append({
            'Method': method_name,
            'Configuration': strategy.description,
            'Towers Located': len(results[method_name]['towers']),
            'Towers Validated': len(errors),
            'Median Error (m)': round(errors.median(), 1),
            'Mean Error (m)': round(errors.mean(), 1),
            'P95 Error (m)': round(errors.quantile(0.95), 1),
            'Within 50m (%)': round((errors <= 50).mean() * 100, 1),
            'Within 100m (%)': round((errors <= 100).mean() * 100, 1),
            'Within 200m (%)': round((errors <= 200).mean() * 100, 1),
            'Within 500m (%)': round((errors <= 500).mean() * 100, 1),
        })

    summary_df = pd.DataFrame(summary_data)

    # ── Sheet 2: Tower-by-Tower Comparison ───────────────────────────────
    # Join all methods on enb_id
    comparison = None
    for i, method_name in enumerate(methods):
        matched = results[method_name]['matched']
        cols_to_use = ['enb_id', 'gt_site_name', 'n_measurements', 'error_m', 'confidence']

        # Only include site name and n_measurements from first method
        if i > 0:
            cols_to_use = ['enb_id', 'error_m', 'confidence']

        method_df = matched[cols_to_use].copy()

        if i == 0:
            comparison = method_df.rename(columns={
                'error_m': f'error_m_{method_name.lower()}',
                'confidence': f'confidence_{method_name.lower()}'
            })
        else:
            method_df = method_df.rename(columns={
                'error_m': f'error_m_{method_name.lower()}',
                'confidence': f'confidence_{method_name.lower()}'
            })
            comparison = comparison.merge(method_df, on='enb_id', how='outer')

    # Calculate which method is better per tower
    error_cols = [f'error_m_{m.lower()}' for m in methods]
    comparison['best_method'] = comparison[error_cols].idxmin(axis=1).str.replace('error_m_', '').str.upper()
    comparison['best_error_m'] = comparison[error_cols].min(axis=1)
    comparison['worst_error_m'] = comparison[error_cols].max(axis=1)
    comparison['error_difference'] = comparison['worst_error_m'] - comparison['best_error_m']

    # Sort by largest improvement
    comparison = comparison.sort_values('error_difference', ascending=False).reset_index(drop=True)

    # ── Build Excel Workbook ─────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    _write_summary_sheet(ws_summary, summary_df, methods)

    # Sheet 2: Tower Comparison
    ws_compare = wb.create_sheet('Tower-by-Tower')
    _write_comparison_sheet(ws_compare, comparison, methods)

    # Sheet 3+: All Errors for each method
    for method_name in methods:
        ws_errors = wb.create_sheet(f'{method_name} All Errors')
        _write_all_errors_sheet(ws_errors, results[method_name]['matched'], method_name)

    wb.save(output_path)
    print(f"  ✅ Comparison report: {output_path}")

    # ── Print summary to console ─────────────────────────────────────────
    print(f"\n  {'─'*60}")
    print(f"  METHOD COMPARISON SUMMARY")
    print(f"  {'─'*60}")

    for _, row in summary_df.iterrows():
        print(f"\n  {row['Method']} Method:")
        print(f"    Configuration:   {row['Configuration']}")
        print(f"    Median error:    {row['Median Error (m)']:.1f} m")
        print(f"    P95 error:       {row['P95 Error (m)']:.1f} m")
        print(f"    Within 200m:     {row['Within 200m (%)']:.1f}%")

    # Winner determination
    median_errors = {row['Method']: row['Median Error (m)'] for _, row in summary_df.iterrows()}
    winner = min(median_errors, key=median_errors.get)
    winner_median = median_errors[winner]
    loser = max(median_errors, key=median_errors.get)
    loser_median = median_errors[loser]
    improvement_pct = ((loser_median - winner_median) / loser_median) * 100

    print(f"\n  🏆 WINNER: {winner}")
    print(f"     Median error {improvement_pct:.1f}% better than {loser}")
    print(f"     ({winner_median:.1f}m vs {loser_median:.1f}m)")
    print(f"  {'─'*60}\n")


def _write_summary_sheet(ws, summary_df, methods):
    """Write statistical summary with formatting."""
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    title_font = Font(name='Arial', bold=True, size=14, color='1E3A5F')
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = 'Positioning Method Comparison — Statistical Summary'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Headers
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = list(summary_df.columns)
    col_widths = [12, 30, 15, 15, 15, 15, 15, 15, 15, 15, 15]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 35

    # Data rows
    data_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)

    for ri, row in enumerate(summary_df.itertuples(index=False), start=4):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal='left' if ci in [1, 2] else 'center',
                vertical='center'
            )
        ws.row_dimensions[ri].height = 20

    # Highlight winner row for key metrics
    median_col = headers.index('Median Error (m)') + 1
    median_values = [ws.cell(row=r, column=median_col).value for r in range(4, 4 + len(summary_df))]
    best_median_row = median_values.index(min(median_values)) + 4

    winner_fill = PatternFill('solid', fgColor='D1FAE5')
    for ci in range(median_col, len(headers) + 1):
        ws.cell(row=best_median_row, column=ci).fill = winner_fill
        ws.cell(row=best_median_row, column=ci).font = bold_font


def _write_comparison_sheet(ws, comparison_df, methods):
    """Write tower-by-tower comparison with color coding."""
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    title_font = Font(name='Arial', bold=True, size=14, color='1E3A5F')
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = 'Tower-by-Tower Comparison'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Prepare display columns
    display_cols = ['enb_id', 'gt_site_name', 'n_measurements']
    display_headers = ['eNBid', 'Site Name', 'Measurements']

    for method in methods:
        display_cols.append(f'error_m_{method.lower()}')
        display_headers.append(f'{method} Error (m)')

    display_cols.extend(['best_method', 'error_difference'])
    display_headers.extend(['Winner', 'Difference (m)'])

    out_df = comparison_df[display_cols].copy()

    # Headers
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    col_widths = [12, 25, 15] + [15] * len(methods) + [12, 15]

    for ci, (h, w) in enumerate(zip(display_headers, col_widths), start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 30

    # Data rows
    data_font = Font(name='Arial', size=10)
    green_fill = PatternFill('solid', fgColor='D1FAE5')
    yellow_fill = PatternFill('solid', fgColor='FEF3C7')

    for ri, row in enumerate(out_df.itertuples(index=False), start=4):
        best_method = row[-2]  # best_method column
        error_diff = row[-1]    # error_difference column

        for ci, val in enumerate(row, start=1):
            # Round floats
            if isinstance(val, (float, np.floating)):
                val = round(val, 1) if not np.isnan(val) else '—'

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal='left' if ci <= 2 else 'center',
                vertical='center'
            )

            # Color-code error columns: green for winner, yellow for loser
            if ci > 3 and ci <= 3 + len(methods):
                method_idx = ci - 4
                method_name = methods[method_idx]
                if method_name == best_method:
                    cell.fill = green_fill
                elif error_diff > 10:  # Only highlight if difference is significant
                    cell.fill = yellow_fill

        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = 'A4'


def _write_all_errors_sheet(ws, matched_df, method_name):
    """Write all tower errors for one method (similar to bs_accuracy.xlsx)."""
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Prepare data
    out = matched_df[['gt_site_name', 'error_m', 'n_measurements']].copy()
    out.columns = ['SITE_NAME', 'Отклонение (м)', 'Кол-во замеров']
    out['Отклонение (м)'] = out['Отклонение (м)'].round(0).astype(int)
    out = out.sort_values('Отклонение (м)').reset_index(drop=True)

    # Header
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_align = Alignment(horizontal='center', vertical='center')

    col_widths = [30, 20, 18]
    for ci, (h, w) in enumerate(zip(out.columns, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    # Data rows with color coding
    data_font = Font(name='Arial', size=10)

    def _row_fill(error_m: float) -> str:
        """Return hex fill color based on location error."""
        if error_m <= 50:  return 'D1FAE5'   # green
        if error_m <= 100: return 'FEF3C7'   # yellow
        if error_m <= 200: return 'FDE8CC'   # orange
        return 'FEE2E2'                      # red

    for ri, row in enumerate(out.itertuples(index=False), start=2):
        site, err, n = row
        fill_color = _row_fill(err)
        fill = PatternFill('solid', fgColor=fill_color)

        for ci, val in enumerate([site, err, n], start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal='left' if ci == 1 else 'center',
                vertical='center'
            )
            if ci in (2, 3):
                cell.fill = fill
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:C{len(out) + 1}'

"""
make_report.py
--------------
Reads merged_towers.csv (output of the validation stage) and produces
a clean Excel file with exactly three columns:

    SITE_NAME  |  Отклонение (м)  |  Кол-во замеров

Rows sorted by error ascending. Color-coded by accuracy tier.

Usage:
    python make_report.py
    python make_report.py --input results/merged_towers.csv --output report.xlsx
"""

import argparse
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Color thresholds ─────────────────────────────────────────────────────────
def _row_fill(error_m: float) -> str:
    """Return hex fill color based on location error."""
    if error_m <= 50:  return 'D1FAE5'   # green
    if error_m <= 100: return 'FEF3C7'   # yellow
    if error_m <= 200: return 'FDE8CC'   # orange
    return 'FEE2E2'                       # red


def make_report(input_path: str = 'merged_towers.csv',
                output_path: str = 'bs_accuracy.xlsx') -> None:

    # ── Load ──────────────────────────────────────────────────────────────
    df = pd.read_csv(input_path)
    print(f"  Loaded {len(df):,} rows from {input_path}")

    # Flexible column detection
    error_col = next((c for c in ['error_m', 'candidate_dist_m'] if c in df.columns), None)
    site_col  = next((c for c in ['gt_site_name', 'site_name']   if c in df.columns), None)
    n_col     = next((c for c in ['n_measurements']               if c in df.columns), None)

    missing = [name for name, col in [('error', error_col), ('site', site_col), ('n', n_col)]
               if col is None]
    if missing:
        sys.exit(f"  ❌ Cannot find columns for: {missing}. "
                 f"Available: {list(df.columns)}")

    out = df[[site_col, error_col, n_col]].copy()
    out.columns = ['SITE_NAME', 'Отклонение (м)', 'Кол-во замеров']
    out['Отклонение (м)'] = out['Отклонение (м)'].round(0).astype(int)
    out = out.sort_values('Отклонение (м)').reset_index(drop=True)

    # ── Print summary ─────────────────────────────────────────────────────
    errs = out['Отклонение (м)']
    print(f"\n  Summary:")
    print(f"    Towers in report:  {len(out):,}")
    print(f"    Median error:      {errs.median():.0f} m")
    print(f"    Within  50m:       {(errs <= 50).mean()*100:.1f}%")
    print(f"    Within 100m:       {(errs <= 100).mean()*100:.1f}%")
    print(f"    Within 200m:       {(errs <= 200).mean()*100:.1f}%")
    print(f"    Within 500m:       {(errs <= 500).mean()*100:.1f}%")

    # ── Build workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Точность БС'

    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', fgColor='1E3A5F')
    hdr_align = Alignment(horizontal='center', vertical='center')

    col_widths = [20, 20, 18]
    for ci, (h, w) in enumerate(zip(out.columns, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    # Data rows
    data_font = Font(name='Arial', size=10)
    for ri, row in enumerate(out.itertuples(index=False), start=2):
        site, err, n = row
        fill = PatternFill('solid', fgColor=_row_fill(err))
        for ci, val in enumerate([site, err, n], start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = data_font
            cell.border    = border
            cell.alignment = Alignment(
                horizontal='left' if ci == 1 else 'center',
                vertical='center',
            )
            if ci in (2, 3):
                cell.fill = fill
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:C{len(out) + 1}'

    wb.save(output_path)
    print(f"\n  ✅ Saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='Generate BS accuracy Excel report')
    parser.add_argument('--input',  default='merged_towers.csv', help='merged_towers.csv path')
    parser.add_argument('--output', default='bs_accuracy.xlsx',  help='Output .xlsx path')
    parser.add_argument('--compare', nargs='+', metavar='CSV',
                        help='Compare multiple methods: --compare merged_towers_ta.csv merged_towers_rsrp.csv')
    args = parser.parse_args()

    print(f"\n{'═'*50}")

    if args.compare:
        # Comparison mode
        if len(args.compare) < 2:
            sys.exit(f"  ❌ --compare requires at least 2 CSV files, got {len(args.compare)}")

        print(f"  METHOD COMPARISON REPORT")
        print(f"{'═'*50}")

        # Load CSVs and build results dict
        from compare_methods import generate_comparison_report
        from distance_strategies import TADistanceStrategy, RSRPDistanceStrategy

        results = {}
        for csv_path in args.compare:
            # Infer method name from filename (e.g., merged_towers_ta.csv → TA)
            if '_ta.' in csv_path or csv_path.endswith('_ta.csv'):
                method_name = 'TA'
                strategy = TADistanceStrategy()
            elif '_rsrp.' in csv_path or csv_path.endswith('_rsrp.csv'):
                method_name = 'RSRP'
                strategy = RSRPDistanceStrategy()
            else:
                # Default: extract from filename
                import re
                match = re.search(r'_(\w+)\.csv$', csv_path)
                method_name = match.group(1).upper() if match else f'Method{len(results)+1}'
                strategy = TADistanceStrategy()  # fallback

            matched_df = pd.read_csv(csv_path)
            results[method_name] = {
                'matched': matched_df,
                'towers': matched_df,  # Approximation (we don't have full towers CSV here)
                'strategy': strategy
            }

        generate_comparison_report(results, args.output)

    else:
        # Single-method mode (original behavior)
        print(f"  BS ACCURACY REPORT")
        print(f"{'═'*50}")
        make_report(args.input, args.output)

    print(f"{'═'*50}\n")


if __name__ == '__main__':
    main()
